import imp
from flask import Flask,render_template,request,flash
from werkzeug.utils import secure_filename 
import os
import cv2
import numpy as np
from tensorflow import keras
from tensorflow.keras.preprocessing import image
from openpyxl.reader.excel import load_workbook
UPLOAD_FOLDER = './files' 



app=Flask(__name__)
dict = {0:'Apple pie',1:'Baby back ribs',2:'Baklava',3:'Beef carpaccio',4:'Beef tartare',5:'Beet salad',6:'Beignets',7:'Bibimbap',8:'Bread pudding',9:'Breakfast burrito',
        10:'Bruschetta',11:'Caesar salad',12:'Cannoli',13:'Caprese salad',14:'Carrot cake',15:'Ceviche',16:'Cheesecake',17:'Cheese plate',18:'Chicken curry',19:'Chicken quesadilla',
        20:'Chicken wings',21:'Chocolate cake',22:'Chocolate mousse',23:'Churros',24:'Clam chowder',25:'Club sandwich',26:'Crab cakes',27:'Creme brulee',28:'Croque madame',
        29:'Cup cakes',30:'Deviled eggs',31:'Donuts',32:'Dumplings',33:'Edamame',34:'Eggs benedict',35:'Escargots',36:'Falafel',37:'Filet mignon',38:'Fish and chips',
        39:'Foie gras',40:'French fries',41:'French onion soup',42:'French toast',43:'Fried calamari',44:'Fried rice',45:'Frozen yogurt',46:'Garlic bread',47:'Gnocchi',
        48:'Greek salad',49:'Grilled cheese sandwich',50:'Grilled salmon',51:'Guacamole',52:'Gyoza',53:'Hamburger',54:'Hot and sour soup',55:'Hot dog',56:'Huevos rancheros',
        57:'Hummus',58:'Ice cream',59:'Lasagna',60:'Lobster bisque',61:'Lobster roll sandwich',62:'Macaroni and cheese',63:'Macarons',64:'Miso soup',65:'Mussels',
        66:'Nachos',67:'Omelette',68:'Onion rings',69:'Oysters',70:'Pad thai',71:'Paella',72:'Pancakes',73:'Panna cotta',74:'Peking duck',75:'Pho',76:'Pizza',77:'Pork chop',
        78:'Poutine',79:'Prime rib',80:'Pulled pork sandwich',81:'Ramen',82:'Ravioli',83:'Red velvet cake',84:'Risotto',85:'Samosa',86:'Sashimi',87:'Scallops',88:'Seaweed salad',
        89:'Shrimp and grits',90:'Spaghetti bolognese',91:'Spaghetti carbonara',92:'Spring rolls',93:'Steak',94:'Strawberry shortcake',95:'Sushi',96:'Tacos',97:'Takoyaki',
        98:'Tiramisu',99:'Tuna tartare',100:'Waffles'}

@app.route('/')
def home():
    return render_template('index.html')
@app.route('/calculate',methods=['POST'])
def cal():
    global height,weight
    height=request.form['height']
    weight=request.form['weight']
    
    BMI=float(weight)/(float(height)**2)
    print(BMI)
    if (BMI < 18.5):
        a="No excess calories found\nBot:under weight"
        return render_template('base.html',pre=a)
    if (BMI >= 18.5 and BMI < 24.9):
        a="Bot: Healthy\nBot: Do cycling 5 mins.. Thats enough"
        return render_template('base.html',pre=a)
    if(BMI  >= 24.9):
       
        a="Bot: Healthy\nBot: Do cycling 5 mins.. Thats enough"
                
    return render_template('base.html',pre=a)
@app.route('/food',methods=['POST'])
def food():
    file=request.files['image']
    basepath = os.path.dirname(__file__)
    file_path = os.path.join(basepath, 'uploads', secure_filename("1.jpg"))
    file.save(file_path)
    model_dl = keras.models.load_model("model.h5") 
    img_to_detect = cv2.imread(r"C:\Users\FABHOSTPYTHON\Desktop\chat bot health care 2\uploads\1.jpg", cv2.IMREAD_COLOR)
    img = cv2.resize(img_to_detect,(64,64))
    x = image.img_to_array(img) 
    x = np.expand_dims(x, axis=0)
    imag = np.vstack([x])
    classes = model_dl.predict_classes(imag,32)
    text = str(dict[classes.item()])
    print(text)
    wrkbk = load_workbook(r"./calorie.xlsx")
    sh = wrkbk.active
    for i in range(1,101):
        c=str(sh.cell(row=i,column=1).value)
      
    if c == text:
        calo =str(sh.cell(row=i,column=2).value)
        a='Bot: The food contain',format(calo),'calories'
        typr =str(sh.cell(row=i,column=4).value)
        print(calo)
        if typr == 'yes':
            bmi =float(weight)/(float(height)**2)
            if (bmi < 18.5):
                # a=Text(window, height=10, width=40,bg="White", fg="black",relief='flat',font=('Times',15,'bold'))
                # a.insert(END,"Bot: No excess calories found\nBot:under weight\nBot:the food contain calories-",text,format(calo))
               
                # a.place(x=700,y=200)
                # break
                print("Bot: No excess calories found\nBot:under weight\nBot:the food contain calories-",text,format(calo))
                  
            elif ( bmi >= 18.5 and bmi < 24.9):
                # a=Text(window, height=10, width=40,bg="White", fg="black",relief='flat',font=('Times',15,'bold'))
                # a.insert(END,"Bot: Healthy\nBot: Do cycling 5 mins.. Thats enough\nthe food contain calories-",c,format(calo))
               
                # a.place(x=700,y=200)
                # #print("Bot: Healthy")
                # #print("Bot: Do cycling 5 mins.. Thats enough")
                # break 
                print("Bot: Healthy\nBot: Do cycling 5 mins.. Thats enough\nthe food contain calories-",c,format(calo))
            elif ( bmi >= 24.9):
                print("Bot:junk food\nthe food contain calories-",c,format(calo))
        elif typr == 'no':
            print("Bot:junk food\nthe food contain calories-",c,format(calo))
    return render_template('result.html',out=text)




if __name__=='__main__':
    app.run(debug=True)